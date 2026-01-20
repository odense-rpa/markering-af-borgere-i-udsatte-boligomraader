
from openpyxl import load_workbook
from typing import List
from pydantic import BaseModel


class BoligData(BaseModel):
    adresse: str
    markering: str


def load_excel_sheet(filepath: str) -> List[BoligData]:
    
    workbook = load_workbook(filepath)
    worksheet = workbook.active
    
    # Get all data excluding headers (first row)
    data_list = []
    
    # Iterate through all rows starting from row 2 (excluding headers)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Get first two columns
        adresse = str(row[0]).strip() if row[0] is not None else ""
        markering = str(row[1]).strip() if row[1] is not None else ""
        
        # Only add if at least one column has data
        if adresse or markering:
            data_list.append(BoligData(adresse=adresse, markering=markering))
    
    return data_list